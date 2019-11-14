import gdxpds
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import os
import re


def xlsdynamicecke(typ, cell, rdim, cdim, sheetname, wb):
    '''
    Returns a list of row and col of bottom-left corner of a table in pandas indexing format (from zero to inf).
    It stops when there is an empty cell in index (rows) or headings (columns).
    typ: string 'set' or 'par'
    cell: string in excel format of top-right table corner cell.
    rdim: indicates the number of columns from the beginning are sets
    cdim: indicates the number of rows from the top are sets
    sheetname: self-explanatory
    wb: is the workbook of an excel file instance of 'from openpyxl import load_workbook'
    eg. xlsdynamicecke('set', C5', 1, 0, 'sheet1', workbook.object)
    return set or table coord.
    '''
    sheet = wb[sheetname]
    string = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'

    def col2num(letters):
        '''
        column letter to column number
        '''
        num = 0
        for c in letters:
            if c in string:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def natural_keys(text):
        '''
        alist.sort(key=natural_keys) sorts in human order
        http://nedbatchelder.com/blog/200712/human_sorting.html
        (See Toothy's implementation in the comments)
        '''
        def atoi(text):
            return int(text) if text.isdigit() else text
        return [atoi(c) for c in re.split(r'(\d+)', text)]

    cut = 0
    for s in cell:
        if s in string:
            cut += 1
        else:
            break
    row = int(cell[cut:])
    col = col2num(cell[:cut])
    if typ == 'par':
        if cdim == 0:
            for i, r in enumerate(sheet.iter_rows(min_row=row+1, min_col=col, max_col=col, values_only=True)):
                if r[0] == None:
                    i = i - 1
                    break
            max_col = rdim + 1
            max_row = row + i
            coord = [row-2, col-1, max_row-1, max_col]
        else:
            for i, c in enumerate(sheet.iter_cols(min_row=row, max_row=row, min_col=col+1, values_only=True)):
                if c[0] == None:
                    i = i - 1
                    break
            max_col = col + i + 1
            for i, r in enumerate(sheet.iter_rows(min_row=row+1, min_col=col, max_col=col, values_only=True)):
                if r[0] == None:
                    i = i - 1
                    break
            max_row = row + i
            coord = [row-1, col-1, max_row-1, max_col]
    elif typ == 'set':
        setls = []
        if rdim == 1:
            for i, r in enumerate(sheet.iter_rows(min_row=row, min_col=col, max_col=col, values_only=True)):
                if r[0] != None:
                    setls.append(r[0])
                else:
                    break
            if all([isinstance(s, (int, float)) for s in list(set(setls))]):
                coord = sorted(list(set(setls)))
            else:
                coord = sorted(list(set(setls)), key=natural_keys)

        elif cdim == 1:
            for i, c in enumerate(sheet.iter_cols(min_row=row, max_row=row, min_col=col, values_only=True)):
                if c[0] != None:
                    setls.append(c[0])
                else:
                    break
            if all([isinstance(s, (int, float)) for s in list(set(setls))]):
                coord = sorted(list(set(setls)))
            else:
                coord = sorted(list(set(setls)), key=natural_keys)
        else:
            raise ValueError('Set must have either rdim or cdim as 1, check dim in py sheet')
    del sheet
    return coord


def exceltogdx(excel_file, gdx_file, csv_file=None):
    '''
    excel_file: input file path
    gdx_file: output file path
    csv_file: if None, it looks at excel file to find sheet with name 'py'
              that constains the instructions to get sets and parameters.
              Otherwise, csv file path.
    '''
    if csv_file == None:
        mapping = pd.read_excel(excel_file, sheet_name='py', index_col='symbol')
    else:
        mapping = pd.read_csv(csv_file, index_col='symbol')

    with open(excel_file, 'rb') as f:
        data = BytesIO(f.read())
    wb = load_workbook(data)
    dc = {}
    for k, v in mapping.iterrows():
        coord = xlsdynamicecke(v['type'], v['startcell'], v['rdim'], v['cdim'], v['sheet_name'], wb)
        if v['type'] == 'par':
            print('par: ', k)
            if v['cdim'] > 0:
                df = pd.read_excel(excel_file, sheet_name=v['sheet_name'], skiprows=coord[0]-1, nrows=coord[2]-coord[0]+2, usecols=range(coord[1], coord[3]))
                # this can be re-code
                os.makedirs('tmp', exist_ok=True)
                df.to_csv(os.path.join('tmp', k+'.csv'), index=False)
                df = pd.read_csv(os.path.join('tmp', k+'.csv'), header=list(range(v['cdim']+1)), skipinitialspace=True)
                # ends prev comment
                df.columns = df.columns.droplevel(0)
                df = df.set_index(df.columns[list(range(v['rdim']))].to_list()).stack(list(range(df.columns.nlevels))).reset_index().rename(columns={0: 'value'})
                dc[k] = df.rename(columns={c: '*' for c in df.columns if c != 'value'})
            else:
                df = pd.read_excel(excel_file, sheet_name=v['sheet_name'], skiprows=coord[0], nrows=coord[2]-coord[0]+1, usecols=range(coord[1], coord[3]))
                df = df.set_index(df.columns[list(range(v['rdim']))].to_list()).rename_axis(['level_0' if v['rdim'] == 1 else None][0], axis=0).reset_index().rename(columns={df.columns.to_list()[-1]: 'value'})
                dc[k] = df.rename(columns={c: '*' for c in df.columns if c != 'value'})
        elif v['type'] == 'set':
            print('set: ', k)
            df = pd.DataFrame({'*': coord})
            df.loc[:, 'value'] = 'c_bool(True)'
            df.dropna(inplace=True)
            dc[k] = df
    print('generating gdx file...')
    gdxpds.to_gdx(dc, gdx_file)
    print('Done!')
    return dc
